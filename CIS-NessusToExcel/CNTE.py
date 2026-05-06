#!/usr/bin/env python3
"""
Convert Nessus .nessus XML compliance / CIS Host Configuration Review scans
into an Excel workbook that is easier to review and report on.

This is a purpose-built Python replacement for the compliance-related parts of
parse_nessus_xml.v24.pl. It focuses on Policy Compliance / CIS benchmark data
rather than attempting to port the entire legacy Perl script 1:1.

Features
--------
- Accepts a single .nessus file or a directory containing multiple .nessus/.xml files
- Extracts Policy Compliance / audit results per host
- Detects benchmark level (L1 / L2) when possible
- Outputs a styled Excel workbook with:
    * Summary
    * All Compliance Checks
    * Host Summary
    * One worksheet per policy/plugin family
- Preserves leading equals signs in values so Excel does not treat them as formulas
- Works well for Nessus Host Configuration Review / CIS benchmark scans

Usage
-----
python nessus_compliance_to_excel.py -f scan.nessus
python nessus_compliance_to_excel.py -d /path/to/scans
python nessus_compliance_to_excel.py -d /path/to/scans -o WDN_CIS_Report.xlsx
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional
import xml.etree.ElementTree as ET

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "Missing dependency: openpyxl. Install it with: pip install openpyxl"
    ) from exc


SUPPORTED_EXTENSIONS = {".nessus", ".xml", ".XML"}
RESULT_ORDER = ["FAILED", "PASSED", "WARNING", "ERROR", "INFO", "UNKNOWN"]


@dataclass
class ComplianceRow:
    source_file: str
    host_name: str
    ip_address: str
    fqdn: str
    operating_system: str
    policy_name: str
    plugin_id: str
    plugin_name: str
    benchmark_level: str
    benchmark_profile: str
    policy_setting: str
    description_of_requirement: str
    solution: str
    result: str
    system_value_or_error: str
    compliance_requirement: str
    severity: str
    service_name: str
    port: str
    protocol: str
    see_also: str
    synopsis: str
    description: str


# ---------- XML helpers ----------

def strip_tag(tag: str) -> str:
    if not tag:
        return ""
    if "}" in tag:
        tag = tag.split("}", 1)[1]
    return tag


def child_text(elem: Optional[ET.Element], name: str, default: str = "") -> str:
    if elem is None:
        return default
    for child in list(elem):
        if strip_tag(child.tag) == name:
            return normalize_text(child.text or "")
    return default


def all_children(elem: Optional[ET.Element], name: str) -> List[ET.Element]:
    if elem is None:
        return []
    return [child for child in list(elem) if strip_tag(child.tag) == name]


def normalize_text(value: Optional[str]) -> str:
    if value is None:
        return ""
    value = value.replace("\r", " ").replace("\n", " ")
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def parse_host_properties(report_host: ET.Element) -> Dict[str, str]:
    props: Dict[str, str] = {}
    host_properties = None
    for child in list(report_host):
        if strip_tag(child.tag) == "HostProperties":
            host_properties = child
            break

    if host_properties is None:
        return props

    for tag in all_children(host_properties, "tag"):
        key = tag.attrib.get("name", "")
        props[key] = normalize_text(tag.text or "")
    return props


# ---------- compliance extraction ----------

def is_compliance_item(report_item: ET.Element) -> bool:
    plugin_family = report_item.attrib.get("pluginFamily", "")
    if plugin_family == "Policy Compliance":
        return True
    names = {strip_tag(c.tag) for c in list(report_item)}
    return any(
        name in names
        for name in {
            "compliance-result",
            "compliance-check-name",
            "compliance-info",
            "compliance-solution",
            "compliance-actual-value",
            "compliance-policy-value",
            "cm_compliance-result",
            "cm_compliance-check-name",
            "cm_compliance-info",
            "cm_compliance-solution",
            "cm_compliance-actual-value",
            "cm_compliance-policy-value",
        }
    )


def extract_benchmark_level(*texts: str) -> str:
    haystack = " ".join(t for t in texts if t).upper()
    if re.search(r"\bL1\b", haystack) or "LEVEL 1" in haystack:
        return "L1"
    if re.search(r"\bL2\b", haystack) or "LEVEL 2" in haystack:
        return "L2"
    return "Unknown"


def extract_benchmark_profile(plugin_name: str, check_name: str, info: str, description: str) -> str:
    haystack = " | ".join([plugin_name, check_name, info, description])
    # Examples seen in some compliance outputs: CIS Microsoft Windows Server 2022 Benchmark
    m = re.search(r"(CIS[^|]{0,140}?Benchmark)", haystack, flags=re.IGNORECASE)
    if m:
        return normalize_text(m.group(1))
    # Fallback to plugin family / name
    return normalize_text(plugin_name) or "Policy Compliance"


def split_policy_setting(check_name: str) -> str:
    # Legacy Perl tried to split on ' - '. We preserve the full check name,
    # but this helper tries to extract the actual setting name when present.
    parts = [p.strip() for p in check_name.split(" - ") if p.strip()]
    if len(parts) >= 2:
        return parts[-1]
    parts = [p.strip() for p in check_name.split(":") if p.strip()]
    if len(parts) >= 2:
        return ": ".join(parts[1:])
    return check_name


def parse_report_item(
    source_file: str,
    host_name: str,
    host_props: Dict[str, str],
    report_item: ET.Element,
) -> ComplianceRow:
    plugin_name = report_item.attrib.get("pluginName", "")
    plugin_id = report_item.attrib.get("pluginID", "")
    severity = report_item.attrib.get("severity", "")
    protocol = report_item.attrib.get("protocol", "")
    port = report_item.attrib.get("port", "")
    service_name = report_item.attrib.get("svc_name", "")

    check_name = child_text(report_item, "cm_compliance-check-name") or child_text(report_item, "compliance-check-name")
    info = child_text(report_item, "cm_compliance-info") or child_text(report_item, "compliance-info")
    solution = child_text(report_item, "cm_compliance-solution") or child_text(report_item, "compliance-solution")
    result = child_text(report_item, "cm_compliance-result") or child_text(report_item, "compliance-result") or "UNKNOWN"
    actual_value = child_text(report_item, "cm_compliance-actual-value") or child_text(report_item, "compliance-actual-value")
    policy_value = child_text(report_item, "cm_compliance-policy-value") or child_text(report_item, "compliance-policy-value")
    see_also = child_text(report_item, "see_also")
    synopsis = child_text(report_item, "synopsis")
    description = child_text(report_item, "description")

    benchmark_level = extract_benchmark_level(plugin_name, check_name, info, description)
    benchmark_profile = extract_benchmark_profile(plugin_name, check_name, info, description)

    return ComplianceRow(
        source_file=source_file,
        host_name=host_name,
        ip_address=host_props.get("host-ip", ""),
        fqdn=host_props.get("host-fqdn", ""),
        operating_system=host_props.get("operating-system", ""),
        policy_name=plugin_name or "Policy Compliance",
        plugin_id=plugin_id,
        plugin_name=plugin_name,
        benchmark_level=benchmark_level,
        benchmark_profile=benchmark_profile,
        policy_setting=split_policy_setting(check_name) if check_name else "",
        description_of_requirement=info,
        solution=solution,
        result=result.upper(),
        system_value_or_error=actual_value,
        compliance_requirement=policy_value,
        severity=severity,
        service_name=service_name,
        port=port,
        protocol=protocol,
        see_also=see_also,
        synopsis=synopsis,
        description=description,
    )


# ---------- file discovery ----------

def find_input_files(single_file: Optional[Path], directory: Optional[Path]) -> List[Path]:
    if bool(single_file) == bool(directory):
        raise ValueError("Use either -f/--file or -d/--directory, but not both.")

    files: List[Path] = []
    if single_file:
        if not single_file.exists() or not single_file.is_file():
            raise FileNotFoundError(f"Input file not found: {single_file}")
        files = [single_file]
    else:
        assert directory is not None
        if not directory.exists() or not directory.is_dir():
            raise FileNotFoundError(f"Input directory not found: {directory}")
        for path in sorted(directory.iterdir()):
            if path.is_file() and path.suffix in SUPPORTED_EXTENSIONS:
                files.append(path)

    if not files:
        raise FileNotFoundError("No .nessus / .xml files were found.")
    return files


# ---------- parser ----------

def _load_xml_root(path: Path) -> ET.Element:
    raw = path.read_text(encoding="utf-8", errors="replace")
    # Nessus compliance data may contain cm: tags without a declared namespace.
    # The legacy Perl parser tolerated this, but ElementTree does not.
    raw = raw.replace("<cm:", "<cm_").replace("</cm:", "</cm_")
    try:
        return ET.fromstring(raw)
    except ET.ParseError as exc:
        raise ValueError(f"XML parse error in {path}: {exc}") from exc


def parse_nessus_file(path: Path) -> List[ComplianceRow]:
    root = _load_xml_root(path)
    if strip_tag(root.tag) != "NessusClientData_v2":
        raise ValueError(f"{path} is not a NessusClientData_v2 file.")

    report = None
    for child in list(root):
        if strip_tag(child.tag) == "Report":
            report = child
            break
    if report is None:
        return []

    rows: List[ComplianceRow] = []
    for report_host in all_children(report, "ReportHost"):
        host_name = report_host.attrib.get("name", "")
        host_props = parse_host_properties(report_host)

        for report_item in all_children(report_host, "ReportItem"):
            if not is_compliance_item(report_item):
                continue
            rows.append(parse_report_item(path.name, host_name, host_props, report_item))
    return rows


# ---------- Excel writing ----------

def safe_sheet_name(name: str, used: set[str]) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", "", name).strip()
    cleaned = cleaned or "Sheet"
    cleaned = cleaned[:31]
    original = cleaned
    counter = 2
    while cleaned in used:
        suffix = f"_{counter}"
        cleaned = f"{original[:31-len(suffix)]}{suffix}"
        counter += 1
    used.add(cleaned)
    return cleaned


def write_dataframe_like(ws, rows: List[Dict[str, str]]) -> None:
    if not rows:
        ws.append(["No data"])
        return

    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        excel_row = []
        for h in headers:
            value = row.get(h, "")
            if isinstance(value, str) and value.startswith("="):
                value = "'" + value
            excel_row.append(value)
        ws.append(excel_row)

    style_worksheet(ws, len(headers))
    add_table(ws, len(rows) + 1, len(headers))


def style_worksheet(ws, num_cols: int) -> None:
    header_fill = PatternFill("solid", fgColor="000000")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="000000")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{ws.max_row}"

    # Widths tuned for review usability
    width_map = {
        "A": 18, "B": 18, "C": 15, "D": 28, "E": 28, "F": 14,
        "G": 40, "H": 14, "I": 25, "J": 60, "K": 60, "L": 18,
        "M": 24, "N": 22, "O": 40, "P": 22, "Q": 10, "R": 16,
        "S": 10, "T": 10, "U": 30, "V": 30, "W": 60,
    }
    for idx in range(1, num_cols + 1):
        col = get_column_letter(idx)
        ws.column_dimensions[col].width = width_map.get(col, 20)


def add_table(ws, max_row: int, max_col: int) -> None:
    if max_row < 2:
        return
    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    table = Table(displayName=f"Table_{re.sub(r'[^A-Za-z0-9]', '_', ws.title)}", ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def build_summary_rows(rows: List[ComplianceRow]) -> List[Dict[str, str]]:
    result_counter = Counter(row.result for row in rows)
    level_counter = Counter(row.benchmark_level for row in rows)
    policy_counter = Counter(row.policy_name for row in rows)
    host_counter = Counter(row.ip_address or row.host_name for row in rows)

    ordered_results = RESULT_ORDER + sorted(set(result_counter) - set(RESULT_ORDER))

    output: List[Dict[str, str]] = []
    output.append({"Metric": "Generated At", "Value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")})
    output.append({"Metric": "Total Compliance Rows", "Value": str(len(rows))})
    output.append({"Metric": "Unique Hosts", "Value": str(len(host_counter))})
    output.append({"Metric": "Unique Policies", "Value": str(len(policy_counter))})
    for result in ordered_results:
        if result in result_counter:
            output.append({"Metric": f"Result - {result}", "Value": str(result_counter[result])})
    for level in sorted(level_counter):
        output.append({"Metric": f"Benchmark Level - {level}", "Value": str(level_counter[level])})
    return output


def build_host_summary_rows(rows: List[ComplianceRow]) -> List[Dict[str, str]]:
    grouped: Dict[str, Counter] = defaultdict(Counter)
    meta: Dict[str, Dict[str, str]] = {}
    for row in rows:
        host_key = row.ip_address or row.host_name
        grouped[host_key][row.result] += 1
        if host_key not in meta:
            meta[host_key] = {
                "Host": row.host_name,
                "IP Address": row.ip_address,
                "FQDN": row.fqdn,
                "Operating System": row.operating_system,
            }

    output: List[Dict[str, str]] = []
    for host_key in sorted(grouped):
        counts = grouped[host_key]
        record = dict(meta[host_key])
        record["FAILED"] = str(counts.get("FAILED", 0))
        record["PASSED"] = str(counts.get("PASSED", 0))
        record["WARNING"] = str(counts.get("WARNING", 0))
        record["ERROR"] = str(counts.get("ERROR", 0))
        record["INFO"] = str(counts.get("INFO", 0))
        record["UNKNOWN"] = str(counts.get("UNKNOWN", 0))
        record["Total"] = str(sum(counts.values()))
        output.append(record)
    return output


def rows_to_dicts(rows: Iterable[ComplianceRow]) -> List[Dict[str, str]]:
    return [asdict(row) for row in rows]


def write_workbook(rows: List[ComplianceRow], output_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    used_sheet_names: set[str] = set()

    summary_ws = wb.create_sheet(safe_sheet_name("Summary", used_sheet_names))
    write_dataframe_like(summary_ws, build_summary_rows(rows))

    all_ws = wb.create_sheet(safe_sheet_name("All Compliance", used_sheet_names))
    write_dataframe_like(all_ws, rows_to_dicts(rows))

    host_ws = wb.create_sheet(safe_sheet_name("Host Summary", used_sheet_names))
    write_dataframe_like(host_ws, build_host_summary_rows(rows))

    by_policy: Dict[str, List[ComplianceRow]] = defaultdict(list)
    for row in rows:
        by_policy[row.policy_name or "Policy Compliance"].append(row)

    for policy_name in sorted(by_policy):
        ws = wb.create_sheet(safe_sheet_name(policy_name, used_sheet_names))
        write_dataframe_like(ws, rows_to_dicts(by_policy[policy_name]))

    wb.save(output_path)


# ---------- CLI ----------

def build_default_output(files: List[Path], single_file: Optional[Path], directory: Optional[Path]) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if single_file:
        return single_file.with_name(f"{single_file.stem}_compliance_{timestamp}.xlsx")
    assert directory is not None
    return directory / f"nessus_compliance_{timestamp}.xlsx"


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description="Convert Nessus Policy Compliance / CIS HCR scans into Excel."
    )
    parser.add_argument("-f", "--file", type=Path, help="Single .nessus/.xml file to parse")
    parser.add_argument("-d", "--directory", type=Path, help="Directory containing .nessus/.xml files")
    parser.add_argument(
        "-o", "--output", type=Path, help="Output .xlsx path (optional)."
    )
    args = parser.parse_args(argv)

    try:
        files = find_input_files(args.file, args.directory)
    except Exception as exc:
        print(f"[!] {exc}", file=sys.stderr)
        return 1

    all_rows: List[ComplianceRow] = []
    for file_path in files:
        print(f"[*] Parsing: {file_path}")
        try:
            rows = parse_nessus_file(file_path)
        except Exception as exc:
            print(f"[!] Failed to parse {file_path}: {exc}", file=sys.stderr)
            return 1
        print(f"    -> extracted {len(rows)} compliance rows")
        all_rows.extend(rows)

    if not all_rows:
        print("[!] No Policy Compliance / CIS compliance rows were found.", file=sys.stderr)
        return 2

    output_path = args.output or build_default_output(files, args.file, args.directory)
    if output_path.suffix.lower() != ".xlsx":
        output_path = output_path.with_suffix(".xlsx")

    write_workbook(all_rows, output_path)
    print(f"[+] Excel workbook written to: {output_path}")
    print(f"[+] Total compliance rows: {len(all_rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
